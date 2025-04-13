import requests
import openpyxl

# Config
TENANT_ID = "f2721b56-1493-4975-8671-51b315109317"
CLIENT_ID = "d23fee23-b5c3-4056-90f3-02244f6f2ab5"
CLIENT_SECRET = "0J38Q~UxwmABLDqsCmurrwCizk2gA7pFW-At7aOe"
TOKEN_URL = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
SCOPE = "Files.Read.All Sites.Read.All User.Read offline_access"
EXCEL_PATH = "E:/Code/Dowload data form one drive to excel/Access_Token.xlsx"

# Load refresh_token từ Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Token Data"]
refresh_token = ws.cell(2, 2).value

# Gửi yêu cầu làm mới token
payload = {
    "client_id": CLIENT_ID,
    "client_secret": CLIENT_SECRET,
    "grant_type": "refresh_token",
    "refresh_token": refresh_token,
    "scope": SCOPE
}

response = requests.post(TOKEN_URL, data=payload)

if response.status_code == 200:
    data = response.json()
    new_access_token = data["access_token"]
    new_refresh_token = data["refresh_token"]

    # Ghi đè lên Excel
    ws.cell(2, 1, new_access_token)
    ws.cell(2, 2, new_refresh_token)
    wb.save(EXCEL_PATH)

    print("✅ Access Token đã được làm mới và lưu lại!")
else:
    print("❌ Lỗi làm mới token:", response.json())
